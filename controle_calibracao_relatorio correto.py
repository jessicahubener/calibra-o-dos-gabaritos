import os
import subprocess
import tkinter as tk
import unicodedata
from datetime import date, datetime, timedelta
from pathlib import Path
from tkinter import filedialog, messagebox, ttk

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


NOME_ARQUIVO_BASE = "DDPRO01_resultado_calibracao"
NOME_ABA_CADASTRO = "Cadastro"
NOME_ABA_DIMENSIONAIS = "Gab. Dimensionais"
NOME_ABA_CERTIFICADOS = "GAB. NORMATIVOS"
NOME_ABA_INSTR_CALIBRADORES = "Instr. Calibradores"
NOME_ABA_PROGRAMACAO = "Programação Diária"
NOME_ABA_PROXIMA_SEMANA = "Cal. Proxima Semana"
NOME_ABA_NAO_ENCONTRADOS = "Gabaritos Não Encontrados"
NOME_ABA_TERCEIROS = "Cal. Terceiros"
NOME_ABA_RESUMO = "Resumo Diário"

LINHA_CABECALHO = 4
COR_NAO_ENCONTRADO = "FF0070C0"
COLUNA_CERTIFICADO = 4
PREFIXO_CODIGO_GABARITO = "GAB-"
MINIMO_GABARITOS_POR_DIA = 4
MINIMO_GABARITOS_TERCEIROS = 15
INTERVALO_ENVIO_TERCEIROS_DIAS = 14
DIA_ENVIO_TERCEIROS = 4

FERIADOS = [
    "01/01/2026",
    "03/04/2026",
    "21/04/2026",
    "01/05/2026",
    "07/09/2026",
    "12/10/2026",
    "02/11/2026",
    "15/11/2026",
    "20/11/2026",
    "25/12/2026",
]

STATUS_VENCIDO = "Vencido"
STATUS_10_DIAS = "Vence em até 10 dias"
STATUS_30_DIAS = "Vence em até 30 dias"
STATUS_PREVENTIVO = "Em dia - programação preventiva"
STATUS_NAO_ENCONTRADO = "Não encontrado para calibração"

PRIORIDADE_VENCIDO = "Prioridade 1"
PRIORIDADE_10_DIAS = "Prioridade 2"
PRIORIDADE_30_DIAS = "Prioridade 3"
PRIORIDADE_PREVENTIVO = "Prioridade 4"
PRIORIDADE_NAO_ENCONTRADO = "Não encontrado"

COR_VENCIDO = "FFFFC7CE"
COR_10_DIAS = "FFD9EAF7"
COR_30_DIAS = "FFFFF2CC"
COR_PREVENTIVO = "FFE2F0D9"
COR_CINZA = "FFD9D9D9"
COR_CABECALHO = "FF1F4E78"
COR_BRANCO = "FFFFFFFF"
CORES_DIAS_SEMANA = {
    "Segunda-feira": "FFEAF3F8",
    "Terça-feira": "FFE2F0D9",
    "Quarta-feira": "FFFFF2CC",
    "Quinta-feira": "FFFCE4D6",
    "Sexta-feira": "FFEDE7F6",
}

PREENCHIMENTO_VENCIDO = PatternFill("solid", fgColor=COR_VENCIDO)
PREENCHIMENTO_10_DIAS = PatternFill("solid", fgColor=COR_10_DIAS)
PREENCHIMENTO_30_DIAS = PatternFill("solid", fgColor=COR_30_DIAS)
PREENCHIMENTO_PREVENTIVO = PatternFill("solid", fgColor=COR_PREVENTIVO)
PREENCHIMENTO_NAO_ENCONTRADO = PatternFill("solid", fgColor=COR_NAO_ENCONTRADO)
PREENCHIMENTO_CABECALHO = PatternFill("solid", fgColor=COR_CABECALHO)

BORDA_FINA = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)
ALINHAMENTO_CENTRO = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALINHAMENTO_ESQUERDA = Alignment(horizontal="left", vertical="center", wrap_text=True)


def converter_data(valor):
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    if isinstance(valor, str):
        texto = valor.strip()
        for formato in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y", "%d-%m-%y"):
            try:
                return datetime.strptime(texto, formato).date()
            except ValueError:
                pass
    return None


def formatar_data(valor):
    data = converter_data(valor)
    if data:
        return data.strftime("%d/%m/%Y")
    if valor in (None, ""):
        return ""
    return str(valor)


def normalizar_texto(valor):
    texto = unicodedata.normalize("NFKD", str(valor or ""))
    texto = "".join(caractere for caractere in texto if not unicodedata.combining(caractere))
    return " ".join(texto.upper().replace(".", " ").replace(",", " ").split())


def normalizar_codigo(valor):
    return str(valor or "").strip().upper()


def codigo_tem_certificado(codigo, codigos_certificados):
    return codigo in codigos_certificados


def valor_indica_sem_certificado(valor):
    texto = normalizar_texto(valor)
    return texto in ("N/A", "NA", "N A", "NAO APLICAVEL", "NÃO APLICÁVEL")


def valor_indica_certificado(valor):
    return valor not in (None, "") and not valor_indica_sem_certificado(valor)


def status_aguardando_calibracao(valor):
    texto = normalizar_texto(valor)
    return "AGUARDANDO" in texto and "CALIBRACAO" in texto


def status_reprovado(valor):
    return normalizar_texto(valor) == "REPROVADO"


def linha_tem_status_aguardando(ws, linha):
    return any(status_aguardando_calibracao(celula.value) for celula in ws[linha])


def linha_tem_status_reprovado(ws, linha):
    return any(status_reprovado(celula.value) for celula in ws[linha])


def registro_deve_ser_ignorado(registro):
    return registro.get("aguardando_calibracao") or registro.get("reprovado")


def encontrar_colunas_por_cabecalho(ws):
    colunas = {}
    for linha in range(1, min(ws.max_row, 10) + 1):
        for celula in ws[linha]:
            texto = normalizar_texto(celula.value)
            if not texto:
                continue
            if texto == "CODIGO":
                colunas["codigo"] = celula.column
            elif texto == "FAMILIA":
                colunas["familia"] = celula.column
            elif texto in ("N", "N°", "Nº"):
                colunas["numero"] = celula.column
            elif texto == "DESCRICAO":
                colunas["descricao"] = celula.column
            elif "SETOR" in texto:
                colunas["setor"] = celula.column
            elif "PROXIMA" in texto and "CALIBRACAO" in texto:
                colunas["proxima_calibracao"] = celula.column
            elif "DATA" in texto and "CALIBRACAO" in texto:
                colunas["data_calibracao"] = celula.column
            elif "PERIODIC" in texto:
                colunas["periodicidade"] = celula.column
    return colunas


def encontrar_codigo_na_linha(ws, linha, coluna_codigo=None):
    if coluna_codigo:
        codigo = normalizar_codigo(ws.cell(linha, coluna_codigo).value)
        if codigo.startswith(PREFIXO_CODIGO_GABARITO):
            return codigo

    for celula in ws[linha]:
        codigo = normalizar_codigo(celula.value)
        if codigo.startswith(PREFIXO_CODIGO_GABARITO):
            return codigo
    return ""


def valor_por_coluna(ws, linha, colunas, nome, padrao=""):
    coluna = colunas.get(nome)
    if not coluna:
        return padrao
    valor = ws.cell(linha, coluna).value
    return padrao if valor in (None, "") else valor


def encontrar_ultima_data_na_linha(ws, linha):
    data_encontrada = None
    for celula in ws[linha]:
        data = converter_data(celula.value)
        if data:
            data_encontrada = data
    return data_encontrada


def localizar_aba_certificados(wb):
    nome_esperado = normalizar_texto(NOME_ABA_CERTIFICADOS)
    for nome_aba in wb.sheetnames:
        if normalizar_texto(nome_aba) == nome_esperado:
            return wb[nome_aba]

    for nome_aba in wb.sheetnames:
        nome = normalizar_texto(nome_aba)
        if "GAB" in nome and "NORMATIVOS" in nome:
            return wb[nome_aba]

    return None


def localizar_aba_instr_calibradores(wb):
    nome_esperado = normalizar_texto(NOME_ABA_INSTR_CALIBRADORES)
    for nome_aba in wb.sheetnames:
        if normalizar_texto(nome_aba) == nome_esperado:
            return wb[nome_aba]

    for nome_aba in wb.sheetnames:
        nome = normalizar_texto(nome_aba)
        if "INSTR" in nome and "CALIBRADORES" in nome:
            return wb[nome_aba]

    return None


def ler_codigos_certificados(wb):
    ws = localizar_aba_certificados(wb)
    if ws is None:
        raise ValueError(
            "A aba GAB. Normativos não foi encontrada."
        )

    codigos = set()
    for linha in range(1, ws.max_row + 1):
        certificado = ws.cell(linha, COLUNA_CERTIFICADO).value
        if not valor_indica_certificado(certificado):
            continue

        for celula in ws[linha]:
            codigo = normalizar_codigo(celula.value)
            if codigo.startswith(PREFIXO_CODIGO_GABARITO):
                codigos.add(codigo)
    return codigos


def feriados_como_datas():
    datas = set()
    for feriado in FERIADOS:
        data = converter_data(feriado)
        if data:
            datas.add(data)
    return datas


def eh_dia_util(data, feriados):
    return data.weekday() < 5 and data not in feriados


def proximo_dia_util(data, feriados):
    while not eh_dia_util(data, feriados):
        data += timedelta(days=1)
    return data


def proxima_data_envio_terceiros(data):
    dias_ate_sexta = (DIA_ENVIO_TERCEIROS - data.weekday()) % 7
    return data + timedelta(days=dias_ate_sexta)


def dias_uteis_da_semana(hoje, feriados):
    inicio = hoje - timedelta(days=hoje.weekday())
    dias = []
    for indice in range(5):
        data = inicio + timedelta(days=indice)
        if eh_dia_util(data, feriados):
            dias.append(data)
    if dias:
        return dias

    data = proximo_dia_util(inicio + timedelta(days=7), feriados)
    while len(dias) < 5:
        if eh_dia_util(data, feriados):
            dias.append(data)
        data += timedelta(days=1)
    return dias


def periodo_da_semana(data):
    inicio = data - timedelta(days=data.weekday())
    fim = inicio + timedelta(days=6)
    return inicio, fim


def pertence_a_semana(proxima_calibracao, hoje):
    data = converter_data(proxima_calibracao)
    if data is None:
        return False

    inicio, fim = periodo_da_semana(hoje)
    return inicio <= data <= fim


def data_posterior_a_semana(proxima_calibracao, semana_referencia):
    data = converter_data(proxima_calibracao)
    if data is None:
        return False

    _, fim = periodo_da_semana(semana_referencia)
    return data > fim


def nome_dia_semana(data):
    nomes = [
        "Segunda-feira",
        "Terça-feira",
        "Quarta-feira",
        "Quinta-feira",
        "Sexta-feira",
        "Sábado",
        "Domingo",
    ]
    return nomes[data.weekday()]


def cor_argb(celula):
    cor = celula.fill.fgColor
    if not cor:
        return ""
    if cor.type == "rgb" and cor.rgb:
        return cor.rgb.upper()
    return ""


def codigo_esta_azul(celula):
    cor = cor_argb(celula)
    return cor in (COR_NAO_ENCONTRADO, COR_NAO_ENCONTRADO[-6:])


def verificar_status(proxima_calibracao, hoje, nao_encontrado=False):
    if nao_encontrado and converter_data(proxima_calibracao) is None:
        return STATUS_NAO_ENCONTRADO, PRIORIDADE_NAO_ENCONTRADO

    data = converter_data(proxima_calibracao)
    if data is None:
        return STATUS_PREVENTIVO, PRIORIDADE_PREVENTIVO

    dias = (data - hoje).days
    if dias < 0:
        return STATUS_VENCIDO, PRIORIDADE_VENCIDO
    if dias <= 10:
        return STATUS_10_DIAS, PRIORIDADE_10_DIAS
    if dias <= 30:
        return STATUS_30_DIAS, PRIORIDADE_30_DIAS
    return STATUS_PREVENTIVO, PRIORIDADE_PREVENTIVO


def linha_principal_completa(registro):
    campos = [
        "codigo",
        "familia",
        "data_calibracao",
        "periodicidade",
        "proxima_calibracao",
        "numero",
        "descricao",
        "setor",
        "status_planilha",
    ]
    return all(registro.get(campo) not in (None, "") for campo in campos)


def chave_prioridade(registro):
    pesos = {
        PRIORIDADE_VENCIDO: 1,
        PRIORIDADE_10_DIAS: 2,
        PRIORIDADE_30_DIAS: 3,
        PRIORIDADE_PREVENTIVO: 4,
    }
    data_calibracao = converter_data(registro["proxima_calibracao"]) or date.max
    return pesos.get(registro["prioridade"], 9), data_calibracao, str(registro["codigo"])


def aplicar_estilo_linha(ws, numero_linha, colunas_esquerda=None):
    colunas_esquerda = set(colunas_esquerda or [])
    for celula in ws[numero_linha]:
        celula.font = Font(size=9)
        celula.border = BORDA_FINA
        celula.alignment = ALINHAMENTO_CENTRO
        if celula.column in colunas_esquerda:
            celula.alignment = ALINHAMENTO_ESQUERDA


def estilizar_cabecalho(ws):
    for celula in ws[1]:
        celula.font = Font(bold=True, color=COR_BRANCO, size=11)
        celula.fill = PREENCHIMENTO_CABECALHO
    aplicar_estilo_linha(ws, 1)
    for celula in ws[1]:
        celula.font = Font(bold=True, color=COR_BRANCO, size=11)
    ws.row_dimensions[1].height = 24


def ajustar_larguras(ws, larguras):
    for coluna, largura in larguras.items():
        ws.column_dimensions[coluna].width = largura


def ajustar_layout_compacto(ws, limites=None, primeira_linha=1):
    limites = limites or {}
    for indice_coluna in range(1, ws.max_column + 1):
        letra = get_column_letter(indice_coluna)
        minimo, maximo = limites.get(letra, (5, 24))
        maior_texto = 0
        for linha in range(primeira_linha, ws.max_row + 1):
            valor = ws.cell(linha, indice_coluna).value
            if valor in (None, ""):
                continue
            partes = str(valor).splitlines() or [""]
            maior_texto = max(maior_texto, *(len(parte) for parte in partes))
        largura = min(maximo, max(minimo, maior_texto + 1))
        ws.column_dimensions[letra].width = largura
    ajustar_alturas_por_texto(ws, primeira_linha=1)


def ajustar_alturas_por_texto(ws, primeira_linha=2):
    for linha in range(primeira_linha, ws.max_row + 1):
        maior_quantidade_linhas = 1
        for celula in ws[linha]:
            texto = str(celula.value or "")
            largura = ws.column_dimensions[celula.column_letter].width or 10
            largura_util = max(1, int(largura * 0.95))
            partes = texto.splitlines() or [""]
            quantidade_linhas = sum(
                max(1, (len(parte) + largura_util - 1) // largura_util)
                for parte in partes
            )
            maior_quantidade_linhas = max(maior_quantidade_linhas, quantidade_linhas)
        altura_minima = 24 if linha == 1 else 16
        ws.row_dimensions[linha].height = max(altura_minima, maior_quantidade_linhas * 13)


def aplicar_cor_status(ws, numero_linha, status):
    preenchimento = None
    if status == STATUS_VENCIDO:
        preenchimento = PREENCHIMENTO_VENCIDO
    elif status == STATUS_10_DIAS:
        preenchimento = PREENCHIMENTO_10_DIAS
    elif status == STATUS_30_DIAS:
        preenchimento = PREENCHIMENTO_30_DIAS
    elif status == STATUS_PREVENTIVO:
        preenchimento = PREENCHIMENTO_PREVENTIVO
    elif status == STATUS_NAO_ENCONTRADO:
        preenchimento = PREENCHIMENTO_NAO_ENCONTRADO

    if preenchimento:
        for celula in ws[numero_linha]:
            celula.fill = preenchimento
            if status == STATUS_NAO_ENCONTRADO:
                celula.font = Font(color=COR_BRANCO, bold=False)


def aplicar_cor_dia_semana(ws, numero_linha, dia_semana, status):
    cor = CORES_DIAS_SEMANA.get(dia_semana)
    if not cor:
        return

    preenchimento = PatternFill("solid", fgColor=cor)
    for celula in ws[numero_linha]:
        celula.fill = preenchimento


def status_para_programacao(registro):
    status = registro["status_calibracao"]
    if registro.get("nao_encontrado"):
        return f"{status} - Não encontrado"
    return status


def definir_arquivo_saida(arquivo_entrada):
    pasta = arquivo_entrada.parent
    base = pasta / f"{NOME_ARQUIVO_BASE}.xlsx"
    if not base.exists():
        return base

    indice = 1
    while True:
        candidato = pasta / f"{NOME_ARQUIVO_BASE}_{indice}.xlsx"
        if not candidato.exists():
            return candidato
        indice += 1


def registrar_gabarito(registros_por_codigo, chave, registro_atual, hoje):
    registro_atual["linha_completa"] = linha_principal_completa(registro_atual)

    if chave in registros_por_codigo:
        registro_existente = registros_por_codigo[chave]
        if registro_atual["nao_encontrado"]:
            registro_existente["nao_encontrado"] = True
            status, prioridade = verificar_status(
                registro_existente["proxima_calibracao"], hoje
            )
            registro_existente["status_calibracao"] = status
            registro_existente["prioridade"] = prioridade
        if (
            not registro_existente.get("linha_completa")
            and registro_atual["linha_completa"]
        ):
            registro_atual["nao_encontrado"] = registro_existente["nao_encontrado"]
            if registro_atual["nao_encontrado"]:
                status, prioridade = verificar_status(
                    registro_atual["proxima_calibracao"], hoje
                )
                registro_atual["status_calibracao"] = status
                registro_atual["prioridade"] = prioridade
            registros_por_codigo[chave] = registro_atual
        return

    registros_por_codigo[chave] = registro_atual


def montar_registro_de_aba_flexivel(ws, linha, colunas, hoje, origem):
    codigo_texto = encontrar_codigo_na_linha(ws, linha, colunas.get("codigo"))
    if not codigo_texto:
        return None

    proxima_calibracao = valor_por_coluna(
        ws,
        linha,
        colunas,
        "proxima_calibracao",
        encontrar_ultima_data_na_linha(ws, linha),
    )
    status_calibracao, prioridade = verificar_status(
        proxima_calibracao, hoje, nao_encontrado=False
    )
    return {
        "familia": valor_por_coluna(ws, linha, colunas, "familia"),
        "codigo": codigo_texto,
        "data_calibracao": valor_por_coluna(ws, linha, colunas, "data_calibracao"),
        "periodicidade": valor_por_coluna(ws, linha, colunas, "periodicidade"),
        "proxima_calibracao": proxima_calibracao,
        "numero": valor_por_coluna(ws, linha, colunas, "numero"),
        "descricao": valor_por_coluna(ws, linha, colunas, "descricao"),
        "setor": valor_por_coluna(ws, linha, colunas, "setor"),
        "status_planilha": "",
        "utilizacao": "",
        "nao_encontrado": False,
        "certificado_terceiros": False,
        "status_calibracao": status_calibracao,
        "prioridade": prioridade,
        "origem": origem,
        "aguardando_calibracao": linha_tem_status_aguardando(ws, linha),
        "reprovado": linha_tem_status_reprovado(ws, linha),
    }


def ler_gabaritos_terceiros(caminho_arquivo, hoje):
    wb = load_workbook(Path(caminho_arquivo), data_only=True)
    abas = []
    ws_normativos = localizar_aba_certificados(wb)
    if ws_normativos is None:
        wb.close()
        raise ValueError("A aba GAB. Normativos não foi encontrada.")
    abas.append((ws_normativos, "GAB. Normativos"))

    ws_instr = localizar_aba_instr_calibradores(wb)
    if ws_instr is None:
        wb.close()
        raise ValueError("A aba Instr. Calibradores não foi encontrada.")
    abas.append((ws_instr, "Instr. Calibradores"))

    registros_por_codigo = {}
    for ws, origem in abas:
        colunas = encontrar_colunas_por_cabecalho(ws)
        for linha in range(1, ws.max_row + 1):
            certificado = ws.cell(linha, COLUNA_CERTIFICADO).value
            if not valor_indica_certificado(certificado):
                continue

            registro = montar_registro_de_aba_flexivel(ws, linha, colunas, hoje, origem)
            if not registro:
                continue
            chave = normalizar_codigo(registro["codigo"])
            registros_por_codigo.setdefault(chave, registro)

    wb.close()
    return list(registros_por_codigo.values())


def datas_envio_terceiros(hoje, feriados):
    fim_ano = date(hoje.year, 12, 31)
    datas = []
    data_base = proxima_data_envio_terceiros(hoje)
    while data_base <= fim_ano:
        data_envio = proximo_dia_util(data_base, feriados)
        if data_envio <= fim_ano and data_envio not in datas:
            datas.append(data_envio)
        data_base += timedelta(days=INTERVALO_ENVIO_TERCEIROS_DIAS)
    return datas


def montar_lote_terceiros(registros, hoje, feriados):
    datas_envio = datas_envio_terceiros(hoje, feriados)
    fim_ano = date(hoje.year, 12, 31)
    com_data = [
        registro
        for registro in registros
        if not registro_deve_ser_ignorado(registro)
        and converter_data(registro["proxima_calibracao"])
        and converter_data(registro["proxima_calibracao"]) <= fim_ano
    ]
    com_data.sort(key=chave_prioridade)

    cronograma = []
    codigos_programados = set()
    for indice, data_envio in enumerate(datas_envio):
        proximo_envio = (
            datas_envio[indice + 1]
            if indice + 1 < len(datas_envio)
            else fim_ano + timedelta(days=1)
        )
        disponiveis = [
            registro
            for registro in com_data
            if normalizar_codigo(registro["codigo"]) not in codigos_programados
        ]
        obrigatorios = [
            registro
            for registro in disponiveis
            if converter_data(registro["proxima_calibracao"]) < proximo_envio
        ]
        codigos_obrigatorios = {
            normalizar_codigo(registro["codigo"]) for registro in obrigatorios
        }
        futuros = [
            registro
            for registro in disponiveis
            if normalizar_codigo(registro["codigo"]) not in codigos_obrigatorios
        ]
        quantidade_completar = max(
            0, MINIMO_GABARITOS_TERCEIROS - len(obrigatorios)
        )
        lote = obrigatorios + futuros[:quantidade_completar]

        for registro in lote:
            item = dict(registro)
            item["data_envio"] = data_envio
            cronograma.append(item)
            codigos_programados.add(normalizar_codigo(registro["codigo"]))

    return cronograma


def ler_gabaritos(caminho_arquivo, hoje):
    arquivo_entrada = Path(caminho_arquivo)
    wb = load_workbook(arquivo_entrada, data_only=True)
    if NOME_ABA_CADASTRO not in wb.sheetnames:
        wb.close()
        raise ValueError(f"A aba '{NOME_ABA_CADASTRO}' não foi encontrada.")
    if NOME_ABA_DIMENSIONAIS not in wb.sheetnames:
        wb.close()
        raise ValueError(f"A aba '{NOME_ABA_DIMENSIONAIS}' não foi encontrada.")

    ws = wb[NOME_ABA_DIMENSIONAIS]
    codigos_certificados = ler_codigos_certificados(wb)
    registros_por_codigo = {}

    for linha in range(LINHA_CABECALHO + 1, ws.max_row + 1):
        codigo = ws.cell(linha, 2).value
        if codigo in (None, ""):
            continue

        codigo_texto = str(codigo).strip()
        chave = normalizar_codigo(codigo_texto)
        nao_encontrado = codigo_esta_azul(ws.cell(linha, 2))
        certificado_terceiros = codigo_tem_certificado(chave, codigos_certificados)
        proxima_calibracao = ws.cell(linha, 5).value
        status_calibracao, prioridade = verificar_status(
            proxima_calibracao, hoje, nao_encontrado
        )
        registro_atual = {
            "familia": ws.cell(linha, 1).value or "",
            "codigo": codigo_texto,
            "data_calibracao": ws.cell(linha, 3).value or "",
            "periodicidade": ws.cell(linha, 4).value or "",
            "proxima_calibracao": proxima_calibracao,
            "numero": ws.cell(linha, 6).value or "",
            "descricao": ws.cell(linha, 7).value or "",
            "setor": ws.cell(linha, 9).value or "",
            "status_planilha": ws.cell(linha, 11).value or "",
            "utilizacao": ws.cell(linha, 12).value or "",
            "nao_encontrado": nao_encontrado,
            "certificado_terceiros": certificado_terceiros,
            "aguardando_calibracao": linha_tem_status_aguardando(ws, linha),
            "reprovado": linha_tem_status_reprovado(ws, linha),
            "status_calibracao": status_calibracao,
            "prioridade": prioridade,
        }
        registro_atual["linha_completa"] = linha_principal_completa(registro_atual)

        if chave in registros_por_codigo:
            registro_existente = registros_por_codigo[chave]
            if codigo_esta_azul(ws.cell(linha, 2)):
                registro_existente["nao_encontrado"] = True
                status, prioridade = verificar_status(
                    registro_existente["proxima_calibracao"], hoje
                )
                registro_existente["status_calibracao"] = status
                registro_existente["prioridade"] = prioridade
            if (
                not registro_existente.get("linha_completa")
                and registro_atual["linha_completa"]
            ):
                registro_atual["nao_encontrado"] = registro_existente["nao_encontrado"]
                if registro_atual["nao_encontrado"]:
                    status, prioridade = verificar_status(
                        registro_atual["proxima_calibracao"], hoje
                    )
                    registro_atual["status_calibracao"] = status
                    registro_atual["prioridade"] = prioridade
                registros_por_codigo[chave] = registro_atual
            continue

        registros_por_codigo[chave] = registro_atual

    ws_cadastro = wb[NOME_ABA_CADASTRO]
    for linha in range(LINHA_CABECALHO + 1, ws_cadastro.max_row + 1):
        codigo = ws_cadastro.cell(linha, 2).value
        if codigo in (None, ""):
            continue

        codigo_texto = str(codigo).strip()
        chave = normalizar_codigo(codigo_texto)
        if codigo_tem_certificado(chave, codigos_certificados):
            continue

        nao_encontrado = codigo_esta_azul(ws_cadastro.cell(linha, 2))
        proxima_calibracao = ws_cadastro.cell(linha, 5).value
        status_calibracao, prioridade = verificar_status(
            proxima_calibracao, hoje, nao_encontrado
        )
        registro_cadastro = {
            "familia": ws_cadastro.cell(linha, 1).value or "",
            "codigo": codigo_texto,
            "data_calibracao": ws_cadastro.cell(linha, 3).value or "",
            "periodicidade": ws_cadastro.cell(linha, 4).value or "",
            "proxima_calibracao": proxima_calibracao,
            "numero": ws_cadastro.cell(linha, 6).value or "",
            "descricao": ws_cadastro.cell(linha, 7).value or "",
            "setor": ws_cadastro.cell(linha, 9).value or "",
            "status_planilha": ws_cadastro.cell(linha, 11).value or "",
            "utilizacao": ws_cadastro.cell(linha, 12).value or "",
            "nao_encontrado": nao_encontrado,
            "certificado_terceiros": False,
            "aguardando_calibracao": linha_tem_status_aguardando(ws_cadastro, linha),
            "reprovado": linha_tem_status_reprovado(ws_cadastro, linha),
            "status_calibracao": status_calibracao,
            "prioridade": prioridade,
        }
        registro_cadastro["linha_completa"] = linha_principal_completa(
            registro_cadastro
        )

        if chave in registros_por_codigo:
            registro_existente = registros_por_codigo[chave]
            if nao_encontrado:
                registro_existente["nao_encontrado"] = True
                registro_existente["status_calibracao"] = status_calibracao
                registro_existente["prioridade"] = prioridade
            if (
                not registro_existente.get("linha_completa")
                and registro_cadastro["linha_completa"]
            ):
                registro_cadastro["nao_encontrado"] = registro_existente["nao_encontrado"]
                if registro_cadastro["nao_encontrado"]:
                    status, prioridade = verificar_status(
                        registro_cadastro["proxima_calibracao"], hoje
                    )
                    registro_cadastro["status_calibracao"] = status
                    registro_cadastro["prioridade"] = prioridade
                registros_por_codigo[chave] = registro_cadastro
            continue

        registros_por_codigo[chave] = registro_cadastro

    ws_normativos = localizar_aba_certificados(wb)
    colunas_normativos = encontrar_colunas_por_cabecalho(ws_normativos)
    for linha in range(1, ws_normativos.max_row + 1):
        certificado = ws_normativos.cell(linha, COLUNA_CERTIFICADO).value
        if not valor_indica_sem_certificado(certificado):
            continue

        codigo_texto = encontrar_codigo_na_linha(
            ws_normativos, linha, colunas_normativos.get("codigo")
        )
        if not codigo_texto:
            continue

        chave = normalizar_codigo(codigo_texto)
        proxima_calibracao = valor_por_coluna(
            ws_normativos,
            linha,
            colunas_normativos,
            "proxima_calibracao",
            encontrar_ultima_data_na_linha(ws_normativos, linha),
        )
        status_calibracao, prioridade = verificar_status(
            proxima_calibracao, hoje, nao_encontrado=False
        )
        registro_normativo = {
            "familia": valor_por_coluna(
                ws_normativos, linha, colunas_normativos, "familia"
            ),
            "codigo": codigo_texto,
            "data_calibracao": valor_por_coluna(
                ws_normativos, linha, colunas_normativos, "data_calibracao"
            ),
            "periodicidade": valor_por_coluna(
                ws_normativos, linha, colunas_normativos, "periodicidade"
            ),
            "proxima_calibracao": proxima_calibracao,
            "numero": valor_por_coluna(
                ws_normativos, linha, colunas_normativos, "numero"
            ),
            "descricao": valor_por_coluna(
                ws_normativos, linha, colunas_normativos, "descricao"
            ),
            "setor": valor_por_coluna(
                ws_normativos, linha, colunas_normativos, "setor"
            ),
            "status_planilha": "",
            "utilizacao": "",
            "nao_encontrado": False,
            "certificado_terceiros": False,
            "aguardando_calibracao": linha_tem_status_aguardando(ws_normativos, linha),
            "reprovado": linha_tem_status_reprovado(ws_normativos, linha),
            "status_calibracao": status_calibracao,
            "prioridade": prioridade,
        }
        registrar_gabarito(registros_por_codigo, chave, registro_normativo, hoje)

    wb.close()
    registros = [
        registro
        for registro in registros_por_codigo.values()
        if not registro.get("certificado_terceiros")
        and not registro.get("reprovado")
    ]
    for registro in registros:
        registro.pop("linha_completa", None)
    return registros


def distribuir_programacao(registros, hoje, feriados, semana_referencia=None):
    semana_referencia = semana_referencia or hoje
    dias_uteis = dias_uteis_da_semana(semana_referencia, feriados)
    programaveis_semana = [
        registro
        for registro in registros
        if not registro_deve_ser_ignorado(registro)
        and pertence_a_semana(registro["proxima_calibracao"], semana_referencia)
    ]
    programaveis_futuros = [
        registro
        for registro in registros
        if not registro_deve_ser_ignorado(registro)
        and data_posterior_a_semana(registro["proxima_calibracao"], semana_referencia)
    ]
    programaveis_semana.sort(key=chave_prioridade)
    programaveis_futuros.sort(key=chave_prioridade)

    meta_minima = len(dias_uteis) * MINIMO_GABARITOS_POR_DIA
    quantidade_adiantar = max(0, meta_minima - len(programaveis_semana))
    programaveis = programaveis_semana + programaveis_futuros[:quantidade_adiantar]
    programaveis.sort(key=chave_prioridade)

    if not dias_uteis:
        return [], 0

    programacao = []
    total = len(programaveis)
    quantidade_base = total // len(dias_uteis)
    sobra = total % len(dias_uteis)
    posicao = 0

    for indice, data_programada in enumerate(dias_uteis):
        quantidade = quantidade_base + (1 if indice < sobra else 0)
        data_programada = proximo_dia_util(data_programada, feriados)
        for registro in programaveis[posicao : posicao + quantidade]:
            item = dict(registro)
            item["data_programada"] = data_programada
            programacao.append(item)
        posicao += quantidade

    media = round(total / len(dias_uteis), 2) if dias_uteis else 0
    return programacao, media


def criar_aba_programacao(wb, programacao, nome_aba=NOME_ABA_PROGRAMACAO):
    if wb.active.max_row == 1 and wb.active.max_column == 1 and wb.active["A1"].value is None:
        ws = wb.active
        ws.title = nome_aba
        ws.delete_rows(1)
    else:
        ws = wb.create_sheet(nome_aba)
    cabecalhos = [
        "Dia da Semana",
        "Data Programada",
        "Código",
        "N°",
        "Descrição",
        "Setor Utilizado",
        "Próxima Calibração",
        "Status da Calibração",
    ]
    ws.append(cabecalhos)
    estilizar_cabecalho(ws)

    for registro in programacao:
        dia_semana = nome_dia_semana(registro["data_programada"])
        ws.append(
            [
                dia_semana,
                formatar_data(registro["data_programada"]),
                registro["codigo"],
                registro["numero"],
                registro["descricao"],
                registro["setor"],
                formatar_data(registro["proxima_calibracao"]),
                status_para_programacao(registro),
            ]
        )
        aplicar_estilo_linha(ws, ws.max_row, colunas_esquerda=(5, 6))
        aplicar_cor_status(ws, ws.max_row, registro["status_calibracao"])
        aplicar_cor_dia_semana(
            ws, ws.max_row, dia_semana, registro["status_calibracao"]
        )

    ajustar_layout_compacto(
        ws,
        {
            "A": (11, 13),
            "B": (10, 11),
            "C": (9, 11),
            "D": (5, 6),
            "E": (24, 44),
            "F": (14, 18),
            "G": (12, 14),
            "H": (16, 19),
        },
    )
    ws.freeze_panes = "A2"


def criar_aba_nao_encontrados(wb, registros, hoje):
    ws = wb.create_sheet(NOME_ABA_NAO_ENCONTRADOS)
    cabecalhos = [
        "Código",
        "N°",
        "Descrição",
        "Setor Utilizado",
        "Próxima Calibração",
        "Status da Calibração",
    ]
    ws.append(cabecalhos)
    estilizar_cabecalho(ws)

    for registro in sorted(
        registros,
        key=lambda item: (
            converter_data(item["proxima_calibracao"]) or date.max,
            str(item["codigo"]),
        ),
    ):
        status_vencimento, _ = verificar_status(
            registro["proxima_calibracao"], hoje, nao_encontrado=False
        )
        ws.append(
            [
                registro["codigo"],
                registro["numero"],
                registro["descricao"],
                registro["setor"],
                formatar_data(registro["proxima_calibracao"]),
                status_vencimento,
            ]
        )
        aplicar_estilo_linha(ws, ws.max_row, colunas_esquerda=(3, 4))
        aplicar_cor_status(ws, ws.max_row, status_vencimento)

    ajustar_layout_compacto(
        ws,
        {
            "A": (9, 11),
            "B": (5, 6),
            "C": (24, 44),
            "D": (14, 18),
            "E": (12, 14),
            "F": (16, 19),
        },
    )
    ws.freeze_panes = "A2"


def criar_aba_terceiros(wb, registros):
    ws = wb.create_sheet(NOME_ABA_TERCEIROS)
    cabecalhos = [
        "Data Envio",
        "Código",
        "N°",
        "Descrição",
        "Setor Utilizado",
        "Próxima Calibração",
        "Status da Calibração",
        "Origem",
    ]
    ws.append(cabecalhos)
    estilizar_cabecalho(ws)

    for registro in registros:
        ws.append(
            [
                formatar_data(registro["data_envio"]),
                registro["codigo"],
                registro["numero"],
                registro["descricao"],
                registro["setor"],
                formatar_data(registro["proxima_calibracao"]),
                registro["status_calibracao"],
                registro.get("origem", ""),
            ]
        )
        aplicar_estilo_linha(ws, ws.max_row, colunas_esquerda=(4, 5, 8))
        aplicar_cor_status(ws, ws.max_row, registro["status_calibracao"])

    ajustar_layout_compacto(
        ws,
        {
            "A": (10, 11),
            "B": (9, 11),
            "C": (5, 6),
            "D": (24, 44),
            "E": (14, 18),
            "F": (12, 14),
            "G": (16, 19),
            "H": (13, 18),
        },
    )
    ws.freeze_panes = "A2"


def criar_aba_resumo(wb, resumo):
    ws = wb.create_sheet(NOME_ABA_RESUMO)
    linhas = [
        ("Data da atualização", formatar_data(resumo["hoje"])),
        (
            "Semana considerada",
            f"{formatar_data(resumo['inicio_semana'])} a {formatar_data(resumo['fim_semana'])}",
        ),
        ("Total de gabaritos analisados", resumo["total_analisados"]),
        ("Total vencidos na semana", resumo["total_vencidos"]),
        ("Total vencendo em até 10 dias na semana", resumo["total_10_dias"]),
        ("Total vencendo em até 30 dias na semana", resumo["total_30_dias"]),
        ("Total em programação preventiva na semana", resumo["total_preventivo"]),
        ("Total não encontrados", resumo["total_nao_encontrados"]),
        ("Quantidade programada para hoje", resumo["programados_hoje"]),
        ("Quantidade programada na semana", resumo["programados_semana"]),
        ("Média sugerida de calibração por dia útil", resumo["media_diaria"]),
        (
            "Sugestão automática",
            f"Para não deixar vencer, calibre aproximadamente "
            f"{resumo['media_diaria']} gabaritos por dia útil.",
        ),
    ]

    ws.append(["Item", "Valor"])
    estilizar_cabecalho(ws)
    for item, valor in linhas:
        ws.append([item, valor])
        aplicar_estilo_linha(ws, ws.max_row)

    ajustar_layout_compacto(ws, {"A": (20, 32), "B": (10, 42)})
    ws.freeze_panes = "A2"


def montar_resumo(registros, programacao, media_diaria, hoje):
    encontrados = [registro for registro in registros if not registro["nao_encontrado"]]
    encontrados_semana = [
        registro
        for registro in encontrados
        if pertence_a_semana(registro["proxima_calibracao"], hoje)
    ]
    nao_encontrados = [registro for registro in registros if registro["nao_encontrado"]]
    inicio_semana, fim_semana = periodo_da_semana(hoje)
    return {
        "hoje": hoje,
        "inicio_semana": inicio_semana,
        "fim_semana": fim_semana,
        "total_analisados": len(registros),
        "total_vencidos": sum(
            1
            for registro in encontrados_semana
            if registro["status_calibracao"] == STATUS_VENCIDO
        ),
        "total_10_dias": sum(
            1
            for registro in encontrados_semana
            if registro["status_calibracao"] == STATUS_10_DIAS
        ),
        "total_30_dias": sum(
            1
            for registro in encontrados_semana
            if registro["status_calibracao"] == STATUS_30_DIAS
        ),
        "total_preventivo": sum(
            1
            for registro in encontrados_semana
            if registro["status_calibracao"] == STATUS_PREVENTIVO
        ),
        "total_nao_encontrados": len(nao_encontrados),
        "programados_hoje": sum(
            1 for registro in programacao if registro["data_programada"] == hoje
        ),
        "programados_semana": len(programacao),
        "media_diaria": media_diaria,
    }


def processar_planilha(caminho_arquivo):
    arquivo_entrada = Path(caminho_arquivo)
    if not arquivo_entrada.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {arquivo_entrada}")

    hoje = datetime.today().date()
    feriados = feriados_como_datas()
    registros = ler_gabaritos(arquivo_entrada, hoje)
    registros_terceiros = ler_gabaritos_terceiros(arquivo_entrada, hoje)
    lote_terceiros = montar_lote_terceiros(registros_terceiros, hoje, feriados)
    programacao, media_diaria = distribuir_programacao(registros, hoje, feriados)
    proxima_semana = hoje + timedelta(days=7)
    programacao_proxima_semana, media_proxima_semana = distribuir_programacao(
        registros, hoje, feriados, semana_referencia=proxima_semana
    )
    for registro in programacao:
        registro["media_diaria"] = media_diaria
    for registro in programacao_proxima_semana:
        registro["media_diaria"] = media_proxima_semana

    nao_encontrados = [registro for registro in registros if registro["nao_encontrado"]]
    resumo = montar_resumo(registros, programacao, media_diaria, hoje)

    arquivo_saida = definir_arquivo_saida(arquivo_entrada)
    wb = Workbook()
    criar_aba_programacao(wb, programacao)
    criar_aba_programacao(wb, programacao_proxima_semana, NOME_ABA_PROXIMA_SEMANA)
    criar_aba_nao_encontrados(wb, nao_encontrados, hoje)
    criar_aba_terceiros(wb, lote_terceiros)
    criar_aba_resumo(wb, resumo)
    wb.save(arquivo_saida)
    wb.close()

    return arquivo_saida, resumo


def abrir_arquivo(caminho):
    if not caminho:
        return
    caminho = str(caminho)
    try:
        os.startfile(caminho)
    except AttributeError:
        subprocess.Popen(["open", caminho])


class AppCalibracao:
    def __init__(self, root):
        self.root = root
        self.root.title("Programação Diária de Calibração")
        self.root.geometry("780x430")
        self.root.resizable(False, False)
        self.caminho_arquivo = tk.StringVar()
        self.arquivo_gerado = None
        self.montar_tela()

    def montar_tela(self):
        frame = ttk.Frame(self.root, padding=20)
        frame.pack(fill="both", expand=True)

        ttk.Label(
            frame,
            text="Programação Diária de Calibração",
            font=("Segoe UI", 16, "bold"),
        ).pack(anchor="w", pady=(0, 8))

        ttk.Label(
            frame,
            text=(
                "Use a planilha DDPRO01. O programa lê a aba Gab. Dimensionais, agrupa os códigos, "
                "identifica não encontrados pela cor azul forte na coluna Código da aba Cadastro "
                "e separa certificados para calibração por terceiros."
            ),
            font=("Segoe UI", 10),
            wraplength=720,
        ).pack(anchor="w", pady=(0, 12))

        linha_arquivo = ttk.Frame(frame)
        linha_arquivo.pack(fill="x", pady=(0, 12))

        ttk.Entry(linha_arquivo, textvariable=self.caminho_arquivo, width=78).pack(
            side="left", fill="x", expand=True
        )
        ttk.Button(
            linha_arquivo,
            text="Selecionar planilha",
            command=self.selecionar_arquivo,
        ).pack(side="left", padx=(8, 0))

        linha_botoes = ttk.Frame(frame)
        linha_botoes.pack(fill="x", pady=(0, 16))

        ttk.Button(
            linha_botoes,
            text="Atualizar programação diária",
            command=self.executar_processamento,
        ).pack(side="left")
        ttk.Button(
            linha_botoes,
            text="Abrir arquivo gerado",
            command=self.abrir_arquivo_gerado,
        ).pack(side="left", padx=(8, 0))
        ttk.Button(linha_botoes, text="Sair", command=self.root.destroy).pack(
            side="left", padx=(8, 0)
        )

        ttk.Label(frame, text="Resultado", font=("Segoe UI", 10, "bold")).pack(
            anchor="w"
        )

        self.caixa_status = tk.Text(
            frame,
            height=14,
            wrap="word",
            font=("Consolas", 10),
            state="disabled",
        )
        self.caixa_status.pack(fill="both", expand=True)
        self.atualizar_status("Selecione a planilha para gerar a programação diária.")

    def atualizar_status(self, mensagem):
        self.caixa_status.config(state="normal")
        self.caixa_status.delete("1.0", tk.END)
        self.caixa_status.insert(tk.END, mensagem)
        self.caixa_status.config(state="disabled")

    def selecionar_arquivo(self):
        caminho = filedialog.askopenfilename(
            title="Selecione a planilha DDPRO01",
            filetypes=[("Arquivos Excel", "*.xlsx")],
        )
        if caminho:
            self.caminho_arquivo.set(caminho)
            self.atualizar_status(f"Arquivo selecionado:\n{caminho}")

    def executar_processamento(self):
        caminho = self.caminho_arquivo.get().strip()
        if not caminho:
            messagebox.showwarning("Aviso", "Selecione uma planilha primeiro.")
            return

        try:
            arquivo_saida, resumo = processar_planilha(caminho)
            self.arquivo_gerado = arquivo_saida
            linhas = [
                "Processamento concluído com sucesso.",
                "",
                f"Arquivo analisado: {caminho}",
                f"Arquivo gerado: {arquivo_saida}",
                "",
                f"Semana considerada: {formatar_data(resumo['inicio_semana'])} a "
                f"{formatar_data(resumo['fim_semana'])}",
                f"Total vencidos na semana: {resumo['total_vencidos']}",
                f"Total vencendo em 10 dias na semana: {resumo['total_10_dias']}",
                f"Total vencendo em 30 dias na semana: {resumo['total_30_dias']}",
                f"Total não encontrados: {resumo['total_nao_encontrados']}",
                f"Quantidade programada para hoje: {resumo['programados_hoje']}",
                f"Quantidade programada na semana: {resumo['programados_semana']}",
                "",
                f"Para não deixar vencer, calibre aproximadamente "
                f"{resumo['media_diaria']} gabaritos por dia útil.",
            ]
            self.atualizar_status("\n".join(linhas))
            messagebox.showinfo("Sucesso", "Programação diária gerada com sucesso.")
        except PermissionError:
            mensagem = (
                "Não foi possível gerar o arquivo. Feche a planilha no Excel "
                "e tente novamente."
            )
            self.atualizar_status(mensagem)
            messagebox.showerror("Erro", mensagem)
        except Exception as erro:
            mensagem = f"Erro ao processar a planilha:\n{erro}"
            self.atualizar_status(mensagem)
            messagebox.showerror("Erro", mensagem)

    def abrir_arquivo_gerado(self):
        if not self.arquivo_gerado:
            messagebox.showwarning("Aviso", "Gere a programação antes de abrir o arquivo.")
            return
        abrir_arquivo(self.arquivo_gerado)


def main():
    root = tk.Tk()
    style = ttk.Style()
    if "vista" in style.theme_names():
        style.theme_use("vista")
    AppCalibracao(root)
    root.mainloop()


if __name__ == "__main__":
    main()
